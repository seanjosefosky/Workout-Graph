import openpyxl, path

from openpyxl import Workbook, load_workbook

path = path.path
wb = load_workbook(path)
ws = wb.active

def Products(wsheet,xlp):
    currentrow = 0
    for row in wsheet.iter_rows(min_col=5, max_col=5):
        currentrow += 1
        for cell in row:
            wsheet[f"E{currentrow}"] = wsheet[f"C{currentrow}"].value * wsheet[f"D{currentrow}"].value

# Products(ws,path)