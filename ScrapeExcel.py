import openpyxl, datetime
from datetime import datetime
from openpyxl import Workbook, load_workbook

wb = load_workbook("workouts.xlsx")
ws = wb.active

# convert date str to datetime format
def date_conversion(string):
    return datetime.strptime(string, "%d %b %Y, %H:%M")

class Exercise():
    def __init__(self):
        self.date = 0
        self.title = 0
        self.weight = 0
        self.reps = 0
        self.vol = 0

# excel rows as objects
exercises = []

# final boss
dict = {} 

# Loop for creating exercises objects
for row in ws.iter_rows(min_row= 2, values_only=False):

    exercise = Exercise()

    for cell in row:
        match cell.column_letter:
            case 'A': continue
            case 'B': exercise.date = date_conversion(cell.value)
            case 'C': continue
            case 'D': continue
            case 'E': exercise.title = cell.value
            case 'F': continue
            case 'G': continue
            case 'H': 
                if cell.value != None:
                    exercise.weight = float(cell.value)
                else:
                    continue
            case 'I': 
                if cell.value != None:
                    exercise.reps = float(cell.value)
                else:
                    continue    
        exercise.vol = exercise.weight * exercise.reps
        exercises.append(exercise)

    if exercise.date not in dict:
        dict[exercise.date] = {}

    if exercise.title not in dict[exercise.date]:
        dict[exercise.date]["exercise"] = exercise.title
        dict[exercise.date]["volume"] = 0


# check if date & title are same & update dict
currentdate = exercises[0].date
currenttitle = exercises[0].title
totalvol = 0

for exercise in exercises:
    if exercise.date == currentdate:
        if exercise.title == currenttitle:
            totalvol += exercise.vol
            dict[exercise.date]["volume"] = f"{totalvol} lbs"
        else: 
            currenttitle = exercise.title
    else:
        currentdate = exercise.date

# for i in dict:
#     print(f"{dict[i]}\n")