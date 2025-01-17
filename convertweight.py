import openpyxl, path

from openpyxl import Workbook, load_workbook

wb = load_workbook(path.path)
ws = wb.active

def ConvertWeight(wsheet,xlp):
        currentrow = 0
        for row in wsheet.iter_rows(min_col=3, max_col=3):
                currentrow += 1
                for cell in row:
                        if cell.value == None:
                                wsheet[f"C{currentrow}"] = 0
                                wb.save(xlp)
                        elif cell.value:
                                wsheet[f"C{currentrow}"] = float(cell.value)
                wb.save(xlp)

# ConvertWeight(ws,path.path)