import openpyxl, path

from openpyxl import Workbook, load_workbook

wb = load_workbook(path.path)
ws = wb.active
wb2 = Workbook()
ws2 = wb2.active

def ConvertNum(ws,ws2):
    for i in range(1, 200):
            ws_cell_value = ws.cell(row=i, column=3).value
            ws2.cell(row=i, column=3).value = float(ws_cell_value)
    wb2.save('test.xlsx')                                       
