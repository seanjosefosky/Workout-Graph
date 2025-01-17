import openpyxl, datetime, path

from datetime import datetime
from openpyxl import Workbook, load_workbook

path = path.path
wb = load_workbook(path)
ws = wb.active

dates = []
title = []
weight = []
reps = []
volumes = []

def ProductSum(ws):
    for row in ws.iter_rows(values_only=False):
        x = 0
        for cell in row:
            match cell.column_letter:
                case 'A': dates.append(cell.value)
                case 'B': title.append(cell.value)
                case 'C': weight.append(float(cell.value))
                case 'D': reps.append(float(cell.value))
        volumes.append(weight[x]*reps[x])
        print(dates[x])
        print(title[x])
        print(f"Weight: {weight[x]}")
        print(f"Reps: {reps[x]}")
        print(f"Volume: {volumes[x]}")  
        x += 1
        


