import openpyxl, path

from openpyxl import Workbook, load_workbook

wb = load_workbook(path.path)
ws = wb.active

def RemoveTreadmill(wsheet,xlp):
        currentrow = 0
        for row in wsheet.iter_rows(min_col=2, max_col=2):
                currentrow += 1
                for cell in row:
                        if cell.value == "Treadmill":
                                ws.delete_rows(currentrow)
                                wb.save(xlp)
                                currentrow -= 1
                                continue
# RemoveTreadmill(ws,path.path)