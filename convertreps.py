import openpyxl, path

from openpyxl import Workbook, load_workbook

wb = load_workbook(path.path)
ws = wb.active

def ConvertReps(wsheet,xlp):
        currentrow = 0
        for row in wsheet.iter_rows(min_col=4, max_col=4):
                currentrow += 1
                for cell in row:
                        if cell.value:
                                wsheet[f"D{currentrow}"] = float(cell.value)
                wb.save(xlp)

# ConvertReps(ws,path.path)