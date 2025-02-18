# Converts Nulls to 0 and strings to floats
import openpyxl, path

from openpyxl import Workbook, load_workbook

wb = load_workbook(path.path)
ws = wb.active

def ConvertToInt(wsheet,xlp):
        currentrow = 0
        for row in wsheet.iter_rows(min_col=4, max_col=4):
                currentrow += 1
                for cell in row:
                        if cell.value:
                                wsheet[f"D{currentrow}"] = float(cell.value)
        currentrow = 0
        for row in wsheet.iter_rows(min_col=3, max_col=3):
                currentrow += 1
                for cell in row:
                        if cell.value == None:
                                wsheet[f"C{currentrow}"] = 0
                        elif cell.value:
                                wsheet[f"C{currentrow}"] = float(cell.value)
        wb.save(xlp)
