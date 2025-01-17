import openpyxl, path
from openpyxl import Workbook, load_workbook
from operator import attrgetter

path = path.path
wb = load_workbook(path)
ws = wb.active

# Initialize a dictionary to store the sums
vol_dict = {}

# Iterate over the rows, starting from row 1 (no headers)
def Consolidate(wsheet,xlp):
    for row in wsheet.iter_rows(min_row=1, max_col=5, values_only=True):
        date = row[0]  # Column A (Date)
        title = row[1]  # Column B (Title)
        value = row[4]  # Column E (Number)
        
        # Use (date, title) tuple as the key
        key = (date, title)
        
        if key in vol_dict:
            vol_dict[key] += value
        else:
            vol_dict[key] = value

    # Print the sum of values for each unique (date, title) pair
    # for key, total in vol_dict.items():
    #     print(f"Date: {key[0]}, Title: {key[1]}, Sum: {total}")

# Consolidate(ws,path)

